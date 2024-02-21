import openpyxl
import re

def find_emails_in_excel(file_path):
    emails = set()

    # Otwórz plik Excel
    wb = openpyxl.load_workbook(file_path)

    # Iteruj przez wszystkie arkusze w pliku Excel
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Iteruj przez wszystkie komórki w arkuszu
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                # Użyj wyrażenia regularnego, aby wyszukać adresy e-mail
                matches = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', str(cell_value))
                emails.update(matches)

    return emails

file_path = 'C:\\Users\\TobiaszFuczek\\Desktop\\Senuku.xlsx'  # Ścieżka do pliku Excel

emails = find_emails_in_excel(file_path)

print("Znalezione adresy e-mail:")
for email in emails:
    print(email)